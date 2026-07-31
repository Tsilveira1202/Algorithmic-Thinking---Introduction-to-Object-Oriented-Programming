"""
Módulo de Orçamento — R.M Imobiliária
=======================================
Define as classes Orcamento, ExportadorCSV e ExportadorExcel.
"""

import csv
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.imovel import Imovel
from models.contrato import Contrato


def formatar_moeda(valor: float) -> str:
    """Formata valor numérico para padrão monetário brasileiro em texto (ex: R$ 1.200,00)."""
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


class Orcamento:
    """
    Gera o orçamento completo de aluguel mensal.

    Combina as informações do imóvel e do contrato para
    produzir um orçamento com 12 parcelas mensais.
    """

    MESES_ORCAMENTO = 12

    def __init__(self, imovel: Imovel, contrato: Contrato, nome_cliente: str = "Cliente"):
        self._imovel = imovel
        self._contrato = contrato
        self._nome_cliente = nome_cliente
        self._data_geracao = datetime.now()

    @property
    def imovel(self) -> Imovel:
        return self._imovel

    @property
    def contrato(self) -> Contrato:
        return self._contrato

    @property
    def nome_cliente(self) -> str:
        return self._nome_cliente

    @property
    def data_geracao(self) -> datetime:
        return self._data_geracao

    def valor_aluguel_mensal(self) -> float:
        return self._imovel.calcular_aluguel()

    def valor_total_12_meses(self) -> float:
        return round(self.valor_aluguel_mensal() * self.MESES_ORCAMENTO, 2)

    def gerar_parcelas(self) -> list:
        aluguel = self.valor_aluguel_mensal()
        num_parcelas_contrato = self._contrato.parcelas
        valor_parcela_contrato = self._contrato.calcular_parcela()

        parcelas = []
        acumulado = 0.0

        for mes in range(1, self.MESES_ORCAMENTO + 1):
            parcela_contrato = valor_parcela_contrato if mes <= num_parcelas_contrato else 0.0
            total_mes = aluguel + parcela_contrato
            acumulado += total_mes

            parcelas.append({
                "mes": mes,
                "descricao": f"Mês {mes:02d}",
                "aluguel": aluguel,
                "parcela_contrato": parcela_contrato,
                "total_mes": total_mes,
                "acumulado": acumulado,
            })

        return parcelas

    def resumo(self) -> dict:
        return {
            "cliente": self._nome_cliente,
            "data": self._data_geracao.strftime("%d/%m/%Y %H:%M"),
            "imovel_tipo": self._imovel.tipo,
            "imovel_quartos": self._imovel.quartos,
            "imovel_garagem": self._imovel.garagem,
            "detalhes_calculo": self._imovel.detalhar_calculo(),
            "aluguel_mensal": self.valor_aluguel_mensal(),
            "total_12_meses": self.valor_total_12_meses(),
            "contrato_valor": self._contrato.valor,
            "contrato_parcelas": self._contrato.parcelas,
            "contrato_valor_parcela": self._contrato.calcular_parcela(),
            "parcelas": self.gerar_parcelas(),
        }

    def __str__(self) -> str:
        aluguel = self.valor_aluguel_mensal()
        return (
            f"Orçamento R.M Imobiliária\n"
            f"Cliente: {self._nome_cliente}\n"
            f"Imóvel: {self._imovel}\n"
            f"Aluguel Mensal: {formatar_moeda(aluguel)}\n"
            f"Total 12 meses: {formatar_moeda(self.valor_total_12_meses())}\n"
            f"Contrato: {self._contrato}"
        )


class ExportadorExcel:
    """
    Gera um relatório profissional em Excel (.xlsx) altamente estilizado,
    com cores corporativas, fontes elegantes, bordas, ajuste de colunas
    e numeração monetária real do Excel.
    """

    @staticmethod
    def exportar(orcamento: Orcamento, caminho: str = None) -> str:
        if caminho is None:
            caminho = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                "orcamento.xlsx"
            )

        resumo = orcamento.resumo()
        parcelas = resumo["parcelas"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orçamento R.M Imobiliária"
        ws.views.sheetView[0].showGridLines = True

        # --- ESTILOS ---
        FONT_FAMILY = "Segoe UI"
        
        # Cores
        COLOR_PRIMARY_DARK = "1A7A5C"    # Verde Escuro Imobiliária
        COLOR_PRIMARY_LIGHT = "EBF5F1"   # Verde Suave Background
        COLOR_ACCENT = "C8944A"          # Âmbar / Dourado
        COLOR_ACCENT_BG = "FFF8EE"       # Âmbar Suave Background
        COLOR_GRAY_BG = "F8F9FA"         # Zebra striping
        COLOR_BORDER = "D1D5DB"          # Cinza claro para bordas

        # Fontes
        font_title = Font(name=FONT_FAMILY, size=15, bold=True, color="FFFFFF")
        font_section = Font(name=FONT_FAMILY, size=12, bold=True, color=COLOR_PRIMARY_DARK)
        font_header = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
        font_bold = Font(name=FONT_FAMILY, size=10, bold=True, color="111827")
        font_regular = Font(name=FONT_FAMILY, size=10, color="1F2937")
        font_total = Font(name=FONT_FAMILY, size=11, bold=True, color="111827")

        # Fills (Preenchimentos)
        fill_title = PatternFill(start_color=COLOR_PRIMARY_DARK, end_color=COLOR_PRIMARY_DARK, fill_type="solid")
        fill_header = PatternFill(start_color=COLOR_PRIMARY_DARK, end_color=COLOR_PRIMARY_DARK, fill_type="solid")
        fill_info_box = PatternFill(start_color=COLOR_PRIMARY_LIGHT, end_color=COLOR_PRIMARY_LIGHT, fill_type="solid")
        fill_total_row = PatternFill(start_color=COLOR_ACCENT_BG, end_color=COLOR_ACCENT_BG, fill_type="solid")
        fill_zebra = PatternFill(start_color=COLOR_GRAY_BG, end_color=COLOR_GRAY_BG, fill_type="solid")

        # Bordas
        thin_side = Side(border_style="thin", color=COLOR_BORDER)
        double_side = Side(border_style="double", color=COLOR_PRIMARY_DARK)
        thick_top_side = Side(border_style="thin", color=COLOR_PRIMARY_DARK)

        border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_total = Border(top=thick_top_side, bottom=double_side, left=thin_side, right=thin_side)

        # Alinhamentos
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # Formato numérico em Excel (PT-BR)
        CURRENCY_FORMAT = 'R$ #,##0.00'

        # --- 1. BANNER DE TÍTULO ---
        ws.merge_cells("A1:F1")
        cell_title = ws["A1"]
        cell_title.value = "R.M IMOBILIÁRIA — ORÇAMENTO DE LOCAÇÃO"
        cell_title.font = font_title
        cell_title.fill = fill_title
        cell_title.alignment = align_center
        ws.row_dimensions[1].height = 40

        # --- 2. QUADRO DE INFORMAÇÕES DO CLIENTE & IMÓVEL ---
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 20
        ws.row_dimensions[5].height = 20

        info_data = [
            ("Cliente:", resumo["cliente"], "Data de Emissão:", resumo["data"]),
            ("Tipo de Imóvel:", resumo["imovel_tipo"], "Quartos:", resumo["imovel_quartos"]),
            ("Garagem / Estacionamento:", "Sim" if resumo["imovel_garagem"] else "Não", "", "")
        ]

        for r_idx, row in enumerate(info_data, start=3):
            ws.cell(row=r_idx, column=1, value=row[0]).font = font_bold
            ws.cell(row=r_idx, column=2, value=row[1]).font = font_regular
            ws.cell(row=r_idx, column=4, value=row[2]).font = font_bold
            ws.cell(row=r_idx, column=5, value=row[3]).font = font_regular

            for c in range(1, 7):
                cell = ws.cell(row=r_idx, column=c)
                cell.fill = fill_info_box
                cell.border = border_all_thin

        # --- 3. SEÇÃO: COMPOSIÇÃO DO ALUGUEL MENSAL ---
        ws.cell(row=7, column=1, value="1. COMPOSIÇÃO DO ALUGUEL MENSAL").font = font_section
        
        headers_comp = ["Item / Descrição", "Tipo", "Valor Mensal (R$)"]
        ws.row_dimensions[8].height = 25
        for col_idx, h in enumerate(headers_comp, start=1):
            cell = ws.cell(row=8, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center if col_idx != 1 else align_left
            cell.border = border_all_thin

        row_idx = 9
        for descricao, valor in resumo["detalhes_calculo"]:
            ws.row_dimensions[row_idx].height = 20
            c1 = ws.cell(row=row_idx, column=1, value=descricao)
            c2 = ws.cell(row=row_idx, column=2, value="Desconto" if valor < 0 else "Base / Adicional")
            c3 = ws.cell(row=row_idx, column=3, value=valor)

            c1.font = font_regular
            c2.font = font_regular
            c3.font = font_regular

            c1.alignment = align_left
            c2.alignment = align_center
            c3.alignment = align_right

            c3.number_format = CURRENCY_FORMAT

            c1.border = border_all_thin
            c2.border = border_all_thin
            c3.border = border_all_thin
            row_idx += 1

        # Total do Aluguel Mensal
        ws.row_dimensions[row_idx].height = 24
        t1 = ws.cell(row=row_idx, column=1, value="ALUGUEL MENSAL TOTAL")
        t2 = ws.cell(row=row_idx, column=2, value="MENSALIDADE")
        t3 = ws.cell(row=row_idx, column=3, value=resumo["aluguel_mensal"])

        t1.font = font_total
        t2.font = font_total
        t3.font = font_total

        t1.alignment = align_left
        t2.alignment = align_center
        t3.alignment = align_right

        t3.number_format = CURRENCY_FORMAT

        for c in (t1, t2, t3):
            c.fill = fill_total_row
            c.border = border_total

        # --- 4. SEÇÃO: CONTRATO IMOBILIÁRIO ---
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="2. CONTRATO IMOBILIÁRIO (TAXA ÚNICA)").font = font_section

        row_idx += 1
        ws.row_dimensions[row_idx].height = 20
        c1 = ws.cell(row=row_idx, column=1, value="Valor Total do Contrato:")
        c2 = ws.cell(row=row_idx, column=2, value=resumo["contrato_valor"])
        c1.font = font_bold
        c2.font = font_bold
        c2.number_format = CURRENCY_FORMAT
        c2.alignment = align_left

        row_idx += 1
        ws.row_dimensions[row_idx].height = 20
        c1 = ws.cell(row=row_idx, column=1, value="Condição de Pagamento:")
        c2 = ws.cell(row=row_idx, column=2, value=f"{resumo['contrato_parcelas']}x de {formatar_moeda(resumo['contrato_valor_parcela'])}")
        c1.font = font_bold
        c2.font = font_regular

        # --- 5. SEÇÃO: TABELA DE 12 PARCELAS (CRONOGRAMA) ---
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="3. CRONOGRAMA FINANCEIRO DE PAGAMENTOS (12 MESES)").font = font_section

        row_idx += 1
        headers_table = ["Mês", "Descrição", "Aluguel Mensal", "Parcela Contrato", "Total Mensal a Pagar", "Total Acumulado"]
        ws.row_dimensions[row_idx].height = 26

        for col_idx, h in enumerate(headers_table, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all_thin

        row_idx += 1
        start_table_row = row_idx

        for p in parcelas:
            ws.row_dimensions[row_idx].height = 20
            c1 = ws.cell(row=row_idx, column=1, value=p["mes"])
            c2 = ws.cell(row=row_idx, column=2, value=p["descricao"])
            c3 = ws.cell(row=row_idx, column=3, value=p["aluguel"])
            c4 = ws.cell(row=row_idx, column=4, value=p["parcela_contrato"])
            c5 = ws.cell(row=row_idx, column=5, value=p["total_mes"])
            c6 = ws.cell(row=row_idx, column=6, value=p["acumulado"])

            c1.alignment = align_center
            c2.alignment = align_center
            c3.alignment = align_right
            c4.alignment = align_right
            c5.alignment = align_right
            c6.alignment = align_right

            for cell in (c1, c2, c3, c4, c5, c6):
                cell.font = font_regular
                cell.border = border_all_thin
                if p["mes"] % 2 == 0:
                    cell.fill = fill_zebra

            c3.number_format = CURRENCY_FORMAT
            c4.number_format = CURRENCY_FORMAT
            c5.number_format = CURRENCY_FORMAT
            c6.number_format = CURRENCY_FORMAT

            row_idx += 1

        # Linha Total Tabela
        ws.row_dimensions[row_idx].height = 25
        end_table_row = row_idx - 1

        t1 = ws.cell(row=row_idx, column=1, value="TOTAL")
        t2 = ws.cell(row=row_idx, column=2, value="12 Meses")
        t3 = ws.cell(row=row_idx, column=3, value=f"=SUM(C{start_table_row}:C{end_table_row})")
        t4 = ws.cell(row=row_idx, column=4, value=f"=SUM(D{start_table_row}:D{end_table_row})")
        t5 = ws.cell(row=row_idx, column=5, value=f"=SUM(E{start_table_row}:E{end_table_row})")
        t6 = ws.cell(row=row_idx, column=6, value=f"=E{end_table_row + 1}")

        t1.alignment = align_center
        t2.alignment = align_center
        t3.alignment = align_right
        t4.alignment = align_right
        t5.alignment = align_right
        t6.alignment = align_right

        for c in (t1, t2, t3, t4, t5, t6):
            c.font = font_total
            c.fill = fill_total_row
            c.border = border_total

        t3.number_format = CURRENCY_FORMAT
        t4.number_format = CURRENCY_FORMAT
        t5.number_format = CURRENCY_FORMAT
        t6.number_format = CURRENCY_FORMAT

        # --- 6. AUTO-FIT NAS COLUNAS ---
        padding = {1: 8, 2: 24, 3: 22, 4: 22, 5: 24, 6: 24}
        for col_idx, width in padding.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        wb.save(caminho)
        return caminho


class ExportadorCSV:
    """
    Exporta o orçamento para um arquivo CSV formatado.
    Mantido para compatibilidade e exportação direta.
    """

    @staticmethod
    def exportar(orcamento: Orcamento, caminho: str = None) -> str:
        if caminho is None:
            caminho = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                "orcamento.csv"
            )

        resumo = orcamento.resumo()
        parcelas = resumo["parcelas"]
        div = "=" * 80

        with open(caminho, mode="w", newline="", encoding="utf-8-sig") as arquivo:
            escritor = csv.writer(arquivo, delimiter=";")

            escritor.writerow(["R.M IMOBILIÁRIA — RELATÓRIO DE ORÇAMENTO DE LOCAÇÃO", "", "", "", "", ""])
            escritor.writerow([div, "", "", "", "", ""])
            escritor.writerow(["Cliente:", resumo["cliente"], "", "Data de Emissão:", resumo["data"], ""])
            escritor.writerow(["Tipo de Imóvel:", resumo["imovel_tipo"], "", "Quartos:", resumo["imovel_quartos"], ""])
            escritor.writerow(["Garagem / Estacionamento:", "Sim" if resumo["imovel_garagem"] else "Não", "", "", "", ""])
            escritor.writerow([])

            escritor.writerow(["1. COMPOSIÇÃO DO ALUGUEL MENSAL", "", "", "", "", ""])
            escritor.writerow(["Item / Descrição", "Tipo", "Valor (R$)", "", "", ""])
            escritor.writerow(["-" * 40, "-" * 15, "-" * 15, "", "", ""])

            for descricao, valor in resumo["detalhes_calculo"]:
                tipo_item = "Desconto" if valor < 0 else "Base/Adicional"
                escritor.writerow([descricao, tipo_item, formatar_moeda(valor), "", "", ""])

            escritor.writerow(["-" * 40, "-" * 15, "-" * 15, "", "", ""])
            escritor.writerow(["VALOR FINAL DO ALUGUEL MENSAL", "MENSALIDADE", formatar_moeda(resumo["aluguel_mensal"]), "", "", ""])
            escritor.writerow([])

            escritor.writerow(["2. CONTRATO IMOBILIÁRIO (TAXA ÚNICA)", "", "", "", "", ""])
            escritor.writerow(["Valor Total do Contrato:", formatar_moeda(resumo["contrato_valor"]), "", "", "", ""])
            escritor.writerow(["Condição de Pagamento:", f"{resumo['contrato_parcelas']}x de {formatar_moeda(resumo['contrato_valor_parcela'])}", "", "", "", ""])
            escritor.writerow([])

            escritor.writerow(["3. CRONOGRAMA FINANCEIRO DE PAGAMENTOS (12 MESES)", "", "", "", "", ""])
            escritor.writerow([
                "Mês",
                "Descrição",
                "Aluguel Mensal (R$)",
                "Parcela Contrato (R$)",
                "Total Mensal (R$)",
                "Total Acumulado (R$)"
            ])
            escritor.writerow(["-" * 6, "-" * 15, "-" * 20, "-" * 20, "-" * 20, "-" * 20])

            total_aluguel = 0.0
            total_contrato = 0.0
            total_geral = 0.0

            for p in parcelas:
                total_aluguel += p["aluguel"]
                total_contrato += p["parcela_contrato"]
                total_geral += p["total_mes"]

                escritor.writerow([
                    f"{p['mes']:02d}",
                    p["descricao"],
                    formatar_moeda(p["aluguel"]),
                    formatar_moeda(p["parcela_contrato"]) if p["parcela_contrato"] > 0 else "R$ 0,00",
                    formatar_moeda(p["total_mes"]),
                    formatar_moeda(p["acumulado"])
                ])

            escritor.writerow(["-" * 6, "-" * 15, "-" * 20, "-" * 20, "-" * 20, "-" * 20])
            escritor.writerow([
                "TOTAL",
                "12 Meses",
                formatar_moeda(total_aluguel),
                formatar_moeda(total_contrato),
                formatar_moeda(total_geral),
                formatar_moeda(total_geral)
            ])
            escritor.writerow([])

            escritor.writerow(["RESUMO GERAL DA LOCAÇÃO", "", "", "", "", ""])
            escritor.writerow(["Total de Aluguel (12 meses):", formatar_moeda(total_aluguel), "", "", "", ""])
            escritor.writerow(["Total do Contrato Imobiliário:", formatar_moeda(total_contrato), "", "", "", ""])
            escritor.writerow(["INVESTIMENTO TOTAL NO PERÍODO (1 ANO):", formatar_moeda(total_geral), "", "", "", ""])
            escritor.writerow([div, "", "", "", "", ""])
            escritor.writerow(["Documento gerado automaticamente pelo Sistema R.M Imobiliária", "", "", "", "", ""])

        return caminho
